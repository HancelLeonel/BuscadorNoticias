import winsound
from googlesearch import search
from openpyxl import Workbook

"""wb=Workbook()"""
wb = Workbook()
ws = wb.active
ws.column_dimensions['A']\
    .width=75
ws.column_dimensions['B']\
    .width=150

nombres = [ ]

for i in range(len(nombres)):
    nombre = nombres[i]
    for url in search(nombre, stop=1):
        ws.row_dimensions[i+1]\
            .height=40
        ws['A'+str(i+1)] = nombre
        ws['B'+str(i+1)] = url
            
    for url2 in search(nombre, stop=2):
       ws['C'+str(i+1)] = url2
    
    for url3 in search(nombre, stop=3):
        ws['D'+str(i+1)] = url3
        print(nombre)
        print(url3)
    
    wb.save('nombres.xlsx')
winsound.PlaySound('SystemExit', winsound.SND_ALIAS)
         





"""for i in range(len(nombres)):
    nombre = nombres[i]
    for url in search(nombre,stop=4):
        print (url)
        noticias = str(url)
        ws.row_dimensions[i+1]\
            .height=50
        ws['A'+str(i+1)] = nombre
        ws['B'+str(i+1)] = noticias
    wb.save('nombres.xlsx')"""
         